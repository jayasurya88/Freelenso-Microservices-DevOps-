from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q
from django.utils import timezone
from django.contrib import messages
from decimal import Decimal
from datetime import timedelta
import json
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

from .models import (
    UserProfile, Project, ProjectApplication, ProjectMilestone, 
    Wallet, Transaction, WithdrawalRequest, PaymentMethod, ProjectActivity, Notification
)

# Helper function to check if user is admin
def is_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    """Main admin dashboard view"""
    # Get counts for dashboard cards
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    total_freelancers = UserProfile.objects.filter(is_freelancer=True).count()
    total_clients = UserProfile.objects.filter(is_client=True).count()
    
    # Project stats
    total_projects = Project.objects.count()
    open_projects = Project.objects.filter(status='open').count()
    in_progress_projects = Project.objects.filter(status='in_progress').count()
    completed_projects = Project.objects.filter(status='completed').count()
    
    # Financial stats
    total_wallet_balance = Wallet.objects.aggregate(total=Sum('balance'))['total'] or 0
    total_escrow_balance = Wallet.objects.aggregate(total=Sum('escrow_balance'))['total'] or 0
    
    # Recent transactions
    recent_transactions = Transaction.objects.order_by('-created_at')[:10]
    
    # Recent projects
    recent_projects = Project.objects.order_by('-created_at')[:5]
    
    # Recent users
    recent_users = User.objects.order_by('-date_joined')[:5]
    
    # Get transaction data for chart (last 7 days)
    chart_data = get_transaction_chart_data()
    
    # Calculate growth rates (last 30 days vs previous 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    sixty_days_ago = timezone.now() - timedelta(days=60)
    
    new_users_current = User.objects.filter(date_joined__gte=thirty_days_ago).count()
    new_users_previous = User.objects.filter(date_joined__gte=sixty_days_ago, date_joined__lt=thirty_days_ago).count()
    new_users_percent = round((new_users_current - new_users_previous) / max(new_users_previous, 1) * 100) if new_users_previous else 100
    
    new_projects_current = Project.objects.filter(created_at__gte=thirty_days_ago).count()
    new_projects_previous = Project.objects.filter(created_at__gte=sixty_days_ago, created_at__lt=thirty_days_ago).count()
    new_projects_percent = round((new_projects_current - new_projects_previous) / max(new_projects_previous, 1) * 100) if new_projects_previous else 100
    
    # Transaction volume growth
    current_volume = Transaction.objects.filter(
        created_at__gte=thirty_days_ago,
        status='completed',
        transaction_type__in=['deposit', 'release']
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    previous_volume = Transaction.objects.filter(
        created_at__gte=sixty_days_ago,
        created_at__lt=thirty_days_ago,
        status='completed',
        transaction_type__in=['deposit', 'release']
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    transaction_volume_growth = round((current_volume - previous_volume) / max(previous_volume, 1) * 100) if previous_volume else 100
    
    # Platform revenue (fees collected)
    current_revenue = Transaction.objects.filter(
        created_at__gte=thirty_days_ago,
        status='completed',
        transaction_type='fee'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    previous_revenue = Transaction.objects.filter(
        created_at__gte=sixty_days_ago,
        created_at__lt=thirty_days_ago,
        status='completed',
        transaction_type='fee'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    revenue_growth = round((current_revenue - previous_revenue) / max(previous_revenue, 1) * 100) if previous_revenue else 100
    
    # Payment processing statistics
    # Razorpay statistics
    razorpay_total = Transaction.objects.filter(
        status='completed',
        description__icontains='razorpay',
        transaction_type='deposit'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Calculate Razorpay success rate
    total_razorpay_attempts = Transaction.objects.filter(
        description__icontains='razorpay',
        transaction_type='deposit'
    ).count()
    
    successful_razorpay = Transaction.objects.filter(
        status='completed',
        description__icontains='razorpay',
        transaction_type='deposit'
    ).count()
    
    razorpay_success_rate = round((successful_razorpay / max(total_razorpay_attempts, 1)) * 100)
    
    # Calculate daily payment processing data (for the chart)
    end_date = timezone.now()
    start_date = end_date - timedelta(days=7)
    
    payment_processing_data = {
        'labels': [],
        'razorpay': [],
        'other_methods': []
    }
    
    # Generate daily payment processing data
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        day_start = current_date.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = current_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Add day label
        payment_processing_data['labels'].append(day_names[current_date.weekday()])
        
        # Razorpay transactions for the day
        razorpay_amount = Transaction.objects.filter(
            transaction_type='deposit',
            status='completed',
            description__icontains='razorpay',
            created_at__range=(day_start, day_end)
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Other payment methods
        other_amount = Transaction.objects.filter(
            transaction_type='deposit',
            status='completed',
            created_at__range=(day_start, day_end)
        ).exclude(description__icontains='razorpay').aggregate(total=Sum('amount'))['total'] or 0
        
        # Add to data
        payment_processing_data['razorpay'].append(float(razorpay_amount))
        payment_processing_data['other_methods'].append(float(other_amount))
    
    # Mock average transaction time (in seconds) since actual processing time might not be stored
    avg_transaction_time = 2.5  # This could be replaced with actual data if available
    
    context = {
        'total_users': total_users,
        'active_users': active_users,
        'total_freelancers': total_freelancers,
        'total_clients': total_clients,
        'total_projects': total_projects,
        'open_projects': open_projects,
        'in_progress_projects': in_progress_projects,
        'completed_projects': completed_projects,
        'total_wallet_balance': total_wallet_balance,
        'total_escrow_balance': total_escrow_balance,
        'recent_transactions': recent_transactions,
        'recent_projects': recent_projects,
        'recent_users': recent_users,
        'chart_data': json.dumps(chart_data),
        'payment_processing_data': json.dumps(payment_processing_data),
        'new_users_percent': new_users_percent,
        'new_projects_percent': new_projects_percent,
        'transaction_volume_growth': transaction_volume_growth,
        'platform_revenue': current_revenue,
        'revenue_growth': revenue_growth,
        'razorpay_total': razorpay_total,
        'razorpay_success_rate': razorpay_success_rate,
        'avg_transaction_time': avg_transaction_time,
        'user_distribution': {
            'freelancers': total_freelancers,
            'clients': total_clients
        }
    }
    
    return render(request, 'admin/dashboard.html', context)

@login_required
@user_passes_test(is_admin)
def admin_users(request):
    """View to manage users"""
    # Get filter parameters
    user_type = request.GET.get('user_type', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    # Base queryset - using UserProfile instead of User
    profiles = UserProfile.objects.all().select_related('user').order_by('-user__date_joined')
    
    # Apply filters
    if user_type == 'freelancer':
        profiles = profiles.filter(is_freelancer=True)
    elif user_type == 'client':
        profiles = profiles.filter(is_client=True)
    elif user_type == 'both':
        profiles = profiles.filter(is_freelancer=True, is_client=True)
    elif user_type == 'admin':
        profiles = profiles.filter(user__is_superuser=True)
        
    if status == 'active':
        profiles = profiles.filter(user__is_active=True)
    elif status == 'inactive':
        profiles = profiles.filter(user__is_active=False)
        
    if search:
        profiles = profiles.filter(
            Q(user__username__icontains=search) | 
            Q(user__email__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search)
        )
    
    # Get some statistics for the view
    total_users = User.objects.count()
    total_freelancers = UserProfile.objects.filter(is_freelancer=True).count()
    total_clients = UserProfile.objects.filter(is_client=True).count()
    new_users_count = User.objects.filter(date_joined__gte=timezone.now() - timedelta(days=30)).count()
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(profiles, 20)  # Show 20 users per page
    
    try:
        profiles = paginator.page(page)
    except PageNotAnInteger:
        profiles = paginator.page(1)
    except EmptyPage:
        profiles = paginator.page(paginator.num_pages)
    
    # Enrich profiles with project counts
    for profile in profiles:
        if profile.is_freelancer:
            profile.completed_projects_count = profile.assigned_projects.filter(status='completed').count()
        else:
            profile.completed_projects_count = 0
            
        if profile.is_client:
            profile.posted_projects_count = profile.posted_projects.count()
        else:
            profile.posted_projects_count = 0
        
        # Calculate total earnings for freelancers
        if profile.is_freelancer:
            # Get sum of all successful transactions for this freelancer
            profile.total_earnings = Transaction.objects.filter(
                wallet__user=profile,
                transaction_type='release',
                status='completed'
            ).aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'profiles': profiles,
        'user_type': user_type,
        'status': status,
        'search': search,
        'total_users': total_users,
        'total_freelancers': total_freelancers,
        'total_clients': total_clients,
        'new_users_count': new_users_count,
    }
    
    return render(request, 'admin/users.html', context)

@login_required
@user_passes_test(is_admin)
def admin_projects(request):
    """View to manage projects"""
    # Get filter parameters
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    # Base queryset
    projects = Project.objects.all().order_by('-created_at')
    
    # Apply filters
    if status:
        projects = projects.filter(status=status)
        
    if search:
        projects = projects.filter(
            Q(title__icontains=search) | 
            Q(description__icontains=search) |
            Q(client__user__username__icontains=search)
        )
    
    context = {
        'projects': projects,
        'status': status,
        'search': search,
        'status_choices': Project.PROJECT_STATUS_CHOICES,
    }
    
    return render(request, 'admin/projects.html', context)

@login_required
@user_passes_test(is_admin)
def admin_transactions(request):
    """View to manage transactions"""
    # Get filter parameters
    transaction_type = request.GET.get('type', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    # Base queryset
    transactions = Transaction.objects.all().order_by('-created_at')
    
    # Apply filters
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
        
    if status:
        transactions = transactions.filter(status=status)
        
    if search:
        transactions = transactions.filter(
            Q(wallet__user__user__username__icontains=search) | 
            Q(transaction_id__icontains=search) |
            Q(description__icontains=search)
        )
    
    context = {
        'transactions': transactions,
        'transaction_type': transaction_type,
        'status': status,
        'search': search,
        'transaction_types': Transaction.TRANSACTION_TYPES,
        'status_choices': Transaction.STATUS_CHOICES,
    }
    
    return render(request, 'admin/transactions.html', context)

@login_required
@user_passes_test(is_admin)
def admin_withdrawals(request):
    """View to manage withdrawal requests"""
    # Get filter parameters
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    # Base queryset
    withdrawals = WithdrawalRequest.objects.all().order_by('-created_at')
    
    # Apply filters
    if status:
        withdrawals = withdrawals.filter(status=status)
        
    if search:
        withdrawals = withdrawals.filter(
            Q(wallet__user__user__username__icontains=search) | 
            Q(reference_id__icontains=search)
        )
    
    context = {
        'withdrawals': withdrawals,
        'status': status,
        'search': search,
        'status_choices': WithdrawalRequest.STATUS_CHOICES,
    }
    
    return render(request, 'admin/withdrawals.html', context)

@login_required
@user_passes_test(is_admin)
def admin_process_withdrawal(request, withdrawal_id):
    """Process a withdrawal request"""
    withdrawal = get_object_or_404(WithdrawalRequest, id=withdrawal_id)
    
    if request.method == 'POST':
        action = request.POST.get('action', '')
        notes = request.POST.get('notes', '')
        reference_id = request.POST.get('reference_id', '')
        
        if action == 'approve':
            withdrawal.status = 'completed'
            withdrawal.processed_at = timezone.now()
            withdrawal.notes = notes
            withdrawal.reference_id = reference_id
            withdrawal.save()
            
            # Update the related transaction
            transaction = Transaction.objects.filter(reference_id=withdrawal.id).first()
            if transaction:
                transaction.status = 'completed'
                transaction.save()
                
            messages.success(request, f'Withdrawal request #{withdrawal.id} has been approved')
            
        elif action == 'reject':
            # Return funds to wallet
            wallet = withdrawal.wallet
            wallet.balance += withdrawal.amount
            wallet.save()
            
            # Update withdrawal status
            withdrawal.status = 'rejected'
            withdrawal.notes = notes
            withdrawal.save()
            
            # Update the related transaction
            transaction = Transaction.objects.filter(reference_id=withdrawal.id).first()
            if transaction:
                transaction.status = 'cancelled'
                transaction.save()
                
            messages.success(request, f'Withdrawal request #{withdrawal.id} has been rejected and funds returned')
            
    return redirect('admin_withdrawals')

@login_required
@user_passes_test(is_admin)
def admin_system_stats(request):
    """View for detailed system statistics"""
    # Time period filter
    period = request.GET.get('period', '30')  # Default to 30 days
    try:
        days = int(period)
        start_date = timezone.now() - timedelta(days=days)
    except ValueError:
        days = 30
        start_date = timezone.now() - timedelta(days=30)
    
    # User stats
    new_users = User.objects.filter(date_joined__gte=start_date).count()
    new_freelancers = UserProfile.objects.filter(
        user__date_joined__gte=start_date, 
        is_freelancer=True
    ).count()
    new_clients = UserProfile.objects.filter(
        user__date_joined__gte=start_date, 
        is_client=True
    ).count()
    
    # Project stats
    new_projects = Project.objects.filter(created_at__gte=start_date).count()
    completed_projects = Project.objects.filter(
        status='completed',
        updated_at__gte=start_date
    ).count()
    
    # Financial stats
    platform_fees = Transaction.objects.filter(
        transaction_type='fee',
        created_at__gte=start_date,
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_transaction_volume = Transaction.objects.filter(
        created_at__gte=start_date,
        status='completed',
        transaction_type__in=['deposit', 'release']
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Monthly growth data
    monthly_users = get_monthly_growth_data(User, 'date_joined')
    monthly_projects = get_monthly_growth_data(Project, 'created_at')
    monthly_transactions = get_monthly_transaction_volume()
    
    context = {
        'days': days,
        'new_users': new_users,
        'new_freelancers': new_freelancers,
        'new_clients': new_clients,
        'new_projects': new_projects,
        'completed_projects': completed_projects,
        'platform_fees': platform_fees,
        'total_transaction_volume': total_transaction_volume,
        'monthly_users': json.dumps(monthly_users),
        'monthly_projects': json.dumps(monthly_projects),
        'monthly_transactions': json.dumps(monthly_transactions),
    }
    
    return render(request, 'admin/system_stats.html', context)

@login_required
@user_passes_test(is_admin)
def admin_cancel_project(request, project_id):
    """Cancel a project and refund any escrowed funds"""
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        # Get the cancellation reason from the form
        cancel_reason = request.POST.get('cancel_reason', 'Project cancelled by administrator')
        
        # Update project status
        original_status = project.status
        project.status = 'cancelled'
        project.save()
        
        # Create project activity entry
        ProjectActivity.objects.create(
            project=project,
            user=request.user,
            activity_type='status_updated',
            description=f"Project cancelled by administrator. Reason: {cancel_reason}"
        )
        
        # Get client wallet
        client_wallet = project.client.wallet
        
        # Get all funded (but not yet approved) milestones
        pending_milestones = project.milestones.filter(status__in=['pending', 'completed'])
        
        total_refunded = Decimal('0.00')
        
        # Process refunds for each milestone
        for milestone in pending_milestones:
            # Check if there are escrow transactions for this milestone
            escrow_transactions = Transaction.objects.filter(
                milestone=milestone,
                transaction_type='escrow',
                status='completed'
            )
            
            if escrow_transactions.exists():
                milestone_amount = milestone.amount
                
                # Refund from escrow to client wallet
                client_wallet.escrow_balance -= milestone_amount
                client_wallet.balance += milestone_amount
                client_wallet.save()
                
                total_refunded += milestone_amount
                
                # Create refund transaction record
                Transaction.objects.create(
                    wallet=client_wallet,
                    project=project,
                    milestone=milestone,
                    amount=milestone_amount,
                    transaction_type='refund',
                    payment_method='wallet',
                    status='completed',
                    description=f"Refund for milestone '{milestone.title}' due to project cancellation"
                )
                
                # Update milestone status
                milestone.status = 'rejected'
                milestone.save()
        
        # Send notifications
        # Notify client
        Notification.objects.create(
            recipient=project.client,
            notification_type='project',
            project=project,
            message=f"Your project '{project.title}' has been cancelled by the administrator. Reason: {cancel_reason}. Any escrowed funds ({total_refunded}) have been returned to your wallet."
        )
        
        # Notify freelancer if assigned
        if project.assigned_freelancer:
            Notification.objects.create(
                recipient=project.assigned_freelancer,
                notification_type='project',
                project=project,
                message=f"The project '{project.title}' has been cancelled by the administrator. Reason: {cancel_reason}."
            )
        
        messages.success(request, f"Project '{project.title}' has been cancelled. {total_refunded} has been refunded to the client's wallet.")
    
    return redirect('admin_projects')

@login_required
@user_passes_test(is_admin)
def admin_export_users(request):
    """Export users data in different formats (CSV, Excel, JSON)"""
    if request.method != 'POST':
        return redirect('admin_users')
    
    # Get form data
    export_format = request.POST.get('format', 'csv')
    user_type = request.POST.get('user_type', '')
    fields = request.POST.getlist('fields', [])
    apply_filters = request.POST.get('apply_filters', False)
    
    # Apply search filters if requested
    if apply_filters:
        current_search = request.POST.get('current_search', '')
        current_user_type = request.POST.get('current_user_type', '')
        current_status = request.POST.get('current_status', '')
    else:
        current_search = ''
        current_user_type = user_type  # Use the export form's user type if not applying search filters
        current_status = ''
    
    # Base queryset - get user profiles instead of users to access related data
    profiles = UserProfile.objects.all().select_related('user')
    
    # Apply filters
    if current_user_type == 'freelancer':
        profiles = profiles.filter(is_freelancer=True)
    elif current_user_type == 'client':
        profiles = profiles.filter(is_client=True)
    elif current_user_type == 'both':
        profiles = profiles.filter(is_freelancer=True, is_client=True)
    elif current_user_type == 'admin':
        profiles = profiles.filter(user__is_superuser=True)
    
    if current_status == 'active':
        profiles = profiles.filter(user__is_active=True)
    elif current_status == 'inactive':
        profiles = profiles.filter(user__is_active=False)
    
    if current_search:
        profiles = profiles.filter(
            Q(user__username__icontains=current_search) | 
            Q(user__email__icontains=current_search) |
            Q(user__first_name__icontains=current_search) |
            Q(user__last_name__icontains=current_search)
        )
    
    # If no fields selected, use defaults
    if not fields:
        fields = ['username', 'email', 'full_name', 'user_type', 'date_joined']
    
    # Prepare the data for export
    export_data = []
    
    # First, determine all possible field names based on the requested fields
    all_field_names = set()
    
    # Pre-define all possible fields that might be added based on user type
    if 'projects' in fields:
        all_field_names.add('Projects Completed')
        all_field_names.add('Projects Posted')
    
    for field in fields:
        if field == 'username':
            all_field_names.add('Username')
        elif field == 'email':
            all_field_names.add('Email')
        elif field == 'full_name':
            all_field_names.add('Full Name')
        elif field == 'user_type':
            all_field_names.add('User Type')
        elif field == 'date_joined':
            all_field_names.add('Date Joined')
        elif field == 'wallet':
            all_field_names.add('Wallet Balance')
    
    for profile in profiles:
        user_data = {field: "" for field in all_field_names}  # Initialize with empty strings
        
        # Include requested fields
        for field in fields:
            if field == 'username':
                user_data['Username'] = profile.user.username
            elif field == 'email':
                user_data['Email'] = profile.user.email
            elif field == 'full_name':
                user_data['Full Name'] = f"{profile.user.first_name} {profile.user.last_name}".strip() or "Not provided"
            elif field == 'user_type':
                user_types = []
                if profile.is_freelancer:
                    user_types.append('Freelancer')
                if profile.is_client:
                    user_types.append('Client')
                if profile.user.is_superuser:
                    user_types.append('Administrator')
                user_data['User Type'] = ', '.join(user_types) or 'User'
            elif field == 'date_joined':
                user_data['Date Joined'] = profile.user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
            elif field == 'projects':
                if profile.is_freelancer:
                    user_data['Projects Completed'] = profile.assigned_projects.filter(status='completed').count()
                if profile.is_client:
                    user_data['Projects Posted'] = profile.posted_projects.count()
            elif field == 'wallet':
                try:
                    wallet = Wallet.objects.get(user=profile)
                    user_data['Wallet Balance'] = wallet.balance
                except Wallet.DoesNotExist:
                    user_data['Wallet Balance'] = 0
        
        export_data.append(user_data)
    
    # Generate response based on format
    if export_format == 'csv':
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="users_export.csv"'
        
        # Write data to CSV
        if export_data:
            writer = csv.DictWriter(response, fieldnames=export_data[0].keys())
            writer.writeheader()
            writer.writerows(export_data)
        
        return response
    
    elif export_format == 'xlsx':
        import openpyxl
        from django.http import HttpResponse
        from io import BytesIO
        
        # Check if openpyxl is installed
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Users"
            
            # Write headers
            if export_data:
                headers = list(export_data[0].keys())
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)
                
                # Write data rows
                for row_num, row_data in enumerate(export_data, 2):
                    for col_num, header in enumerate(headers, 1):
                        ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
            
            # Save to buffer and return response
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
            return response
        
        except ImportError:
            messages.error(request, "Excel export requires the openpyxl package. Please install it or choose another format.")
            return redirect('admin_users')
    
    elif export_format == 'json':
        from django.http import JsonResponse
        
        return JsonResponse({'users': export_data})
    
    # Default fallback
    return redirect('admin_users')

# Helper functions for charts and data analysis
def get_transaction_chart_data():
    """Get transaction data for the last 7 days for charts"""
    end_date = timezone.now()
    start_date = end_date - timedelta(days=7)
    
    # Initialize data structure
    data = {
        'labels': [],
        'deposits': [],
        'withdrawals': [],
        'escrow': [],
        'releases': []
    }
    
    # Generate data for each day
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        date_label = current_date.strftime('%b %d')
        data['labels'].append(date_label)
        
        # Get transaction amounts for the day
        day_start = current_date.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = current_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        deposits = Transaction.objects.filter(
            transaction_type='deposit',
            status='completed',
            created_at__range=(day_start, day_end)
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        withdrawals = Transaction.objects.filter(
            transaction_type='withdrawal',
            status='completed',
            created_at__range=(day_start, day_end)
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        escrow = Transaction.objects.filter(
            transaction_type='escrow',
            status='completed',
            created_at__range=(day_start, day_end)
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        releases = Transaction.objects.filter(
            transaction_type='release',
            status='completed',
            created_at__range=(day_start, day_end)
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        data['deposits'].append(float(deposits))
        data['withdrawals'].append(float(withdrawals))
        data['escrow'].append(float(escrow))
        data['releases'].append(float(releases))
    
    return data

def get_monthly_growth_data(model, date_field):
    """Get monthly growth data for a model"""
    end_date = timezone.now()
    start_date = end_date - timedelta(days=365)  # Last 12 months
    
    data = {
        'labels': [],
        'counts': []
    }
    
    # Get data for each month
    for i in range(12):
        month_start = (end_date - timedelta(days=30 * i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if i < 11:
            month_end = (end_date - timedelta(days=30 * (i+1))).replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(microseconds=1)
        else:
            month_end = start_date
        
        # Skip if month_end is in the future
        if month_end > timezone.now():
            continue
            
        filter_kwargs = {
            f'{date_field}__range': (month_end, month_start)
        }
        
        count = model.objects.filter(**filter_kwargs).count()
        
        month_label = month_start.strftime('%b %Y')
        data['labels'].insert(0, month_label)
        data['counts'].insert(0, count)
    
    return data

def get_monthly_transaction_volume():
    """Get monthly transaction volume data"""
    end_date = timezone.now()
    start_date = end_date - timedelta(days=365)  # Last 12 months
    
    data = {
        'labels': [],
        'amounts': []
    }
    
    # Get data for each month
    for i in range(12):
        month_start = (end_date - timedelta(days=30 * i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if i < 11:
            month_end = (end_date - timedelta(days=30 * (i+1))).replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(microseconds=1)
        else:
            month_end = start_date
        
        # Skip if month_end is in the future
        if month_end > timezone.now():
            continue
            
        amount = Transaction.objects.filter(
            created_at__range=(month_end, month_start),
            status='completed',
            transaction_type__in=['deposit', 'release']
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        month_label = month_start.strftime('%b %Y')
        data['labels'].insert(0, month_label)
        data['amounts'].insert(0, float(amount))
    
    return data

@login_required
@user_passes_test(is_admin)
def admin_activate_user(request, user_id):
    """Activate a user account"""
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        user.is_active = True
        user.save()
        
        messages.success(request, f"User '{user.username}' has been activated successfully.")
        
        # Create notification for the user
        try:
            profile = UserProfile.objects.get(user=user)
            Notification.objects.create(
                recipient=profile,
                notification_type='account',
                message="Your account has been activated by an administrator. You can now log in and use all platform features."
            )
        except UserProfile.DoesNotExist:
            pass  # No profile to notify
    
    return redirect('admin_users')

@login_required
@user_passes_test(is_admin)
def admin_deactivate_user(request, user_id):
    """Deactivate a user account"""
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        
        # Don't allow deactivating yourself
        if user == request.user:
            messages.error(request, "You cannot deactivate your own admin account.")
            return redirect('admin_users')
        
        user.is_active = False
        user.save()
        
        messages.success(request, f"User '{user.username}' has been deactivated successfully.")
        
        # Create notification for the user
        try:
            profile = UserProfile.objects.get(user=user)
            Notification.objects.create(
                recipient=profile,
                notification_type='account',
                message="Your account has been deactivated by an administrator. Please contact support if you believe this was done in error."
            )
        except UserProfile.DoesNotExist:
            pass  # No profile to notify
    
    return redirect('admin_users') 